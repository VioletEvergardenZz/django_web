from django.shortcuts import render,redirect
from app01 import models
from app01.utils.pagination import Pagination
# Create your views here.

'部门列表'
def depart_list(request):
    # 去数据库中获取所有的部门列表
    # [对象,对象,对象]
    queryset = models.Department.objects.all()
    page_object = Pagination(request,queryset,page_size=10)
    context = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html(),
    }
    return render(request,'depart_list.html',context)

'添加部门'
def depart_add(request):
    if request.method == "GET":
        return render(request,'depart_add.html')

    # 获取用户POST提交过来的数据
    title = request.POST.get('title')
    # 保存到数据库
    models.Department.objects.create(title=title)
    # 重定向回部门列表
    return redirect('/depart/list/')

'删除部门'
def depart_delete(request):
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect('/depart/list/')

'修改部门'
def depart_edit(request,nid):
    if request.method == "GET":
        # 根据nid 获取它的数据
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request,'depart_edit.html',{'row_object':row_object})

    # 获取用户提交的标题
    title = request.POST.get('title')
    # 根据id 找到数据库中的数据并进行更新
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect('/depart/list/')

'批量删除（Excel文件）'
def depart_multi(request):
    from openpyxl import load_workbook
    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")

    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect('/depart/list/')